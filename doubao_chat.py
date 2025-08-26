# -*- coding: utf-8 -*-
import uuid
import time
import json
import pandas as pd
import datetime
import os
import math
from get_doubao_result import call_doubao_result
from get_doubao_result import clear_chat
from playwright.sync_api import sync_playwright
from get_chatgpt_result import get_chatgpt_result
from format_tool.json_tool import *
base_path = os.path.expanduser('')

history_map_full = {}
history_map_recent = {}
global id_sum
id_sum=0
from datetime import datetime

def get_current_time():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def read_excel(file_path):
    sheets = pd.read_excel(file_path, sheet_name=None)
    return sheets

def generate_gpt_prompt(template, yp_name, player_prompt_from_player, user_nameprompt, history_str):
    prompt = template.format(yp_name=yp_name, player=player_prompt_from_player, user_nameprompt=user_nameprompt, history_str=history_str)
    return prompt


def call_gpt_api(prompt):
    query = get_chatgpt_result(prompt)
    return query

def update_history_full(yp_name,current_qa_dict):
    global history_map_full
    if yp_name not in history_map_full:
        history_map_full[yp_name] = []
    history_map_full[yp_name].append(current_qa_dict)

def update_history_recent(yp_name,current_qa_dict):

    global history_map_recent
    if yp_name not in history_map_recent:
        history_map_recent[yp_name] = []
    history_map_recent[yp_name].append(current_qa_dict)

def get_recent_history_str(yp_name, num):

    global history_map_recent
    if yp_name not in history_map_recent:
        return ""
    recent = history_map_recent[yp_name][-num:]
    history_str = "\n".join([f"{entry['current_qa']}" for entry in recent])
    return history_str

def get_recent_history_str_full(yp_name, num):
    global history_map_full
    if yp_name not in history_map_full:
        return ""
    recent = history_map_full[yp_name][-num:]
    history_str = "\n".join([f"{entry['current_qa']}" for entry in recent])
    return history_str

def get_full_history_str(yp_name):
    global history_map_full
    if yp_name not in history_map_full:
        return ""
    full = history_map_full[yp_name]
    full_str = "\n".join([f"{entry['current_qa']}" for entry in full])
    return full_str

def append_jsonl(jsonl_file_path, json_object):
    with open(jsonl_file_path, "a", encoding="utf-8") as f:
        f.write(json.dumps(json_object, ensure_ascii=False) + "\n")

def doubao_star(playwright,url,state_file,headless):
    c=0
    while True:
        try:
            browser = playwright.chromium.launch(headless=headless)
            context = browser.new_context(storage_state=state_file)
            page = context.new_page()
            page.goto(url)
            time.sleep(3.0)
            print("打开浏览器成功...")
            return page,context,browser
        except Exception:
            c+=1
            pass
        if c > 10:
            print("打开浏览器出错10次...")
            break

def doubao_end(page,context,browser):
    page.close()
    context.close()
    browser.close()

def handle_conversation_row(row, gpt_prompt_template, player_df, jsonl_file_path,yp_name,index,page,numbers):
    一级分类 = row["一级分类"]
    二级分类 = row["二级分类"]
    三级分类 = row["三级分类"]
    对话形式 = row["对话形式"]
    user_nameprompt = row["user_nameprompt"]
    序号 = row["序号"]
    yp_name1 = row["multiypname"]
    ypfixedfirst_sentence = row["ypfixedfirst_sentence"] if "ypfixedfirst_sentence" in row and not pd.isna(row["ypfixedfirst_sentence"]) else ""
    user_namefixedquestion = row["user_namefixedquestion"] if "user_namefixedquestion" in row and not pd.isna(row["user_namefixedquestion"]) else ""
    user_nametype1 = row["user_nametype1"] if "user_nametype1" in row and not pd.isna(row["user_nametype1"]) else ""

    if 对话形式 == "ypfirst_sentencefixed":
        first_sentence = str(ypfixedfirst_sentence)
        initiator = "yp"

    elif 对话形式 == "user_name发起，user_namefirst_sentencefixed":
        first_sentence = str(user_namefixedquestion)
        initiator = "user_name"
    else:
        return
    print("发起question的一方为:",initiator)

    allowed_types = [s.strip() for s in str(user_nametype1).split("\n") if s.strip() != ""]
    applicable_players = player_df[player_df["player_type"].isin(allowed_types)]
    print("user_nametype1:", applicable_players)

    first_staus=True

    print(first_sentence)
    aaa = True
    对话轮次 = row["对话轮次"]
    qa_num = int(对话轮次)
    print("对话轮次", qa_num)
    for round_num in range(1, qa_num+1):
        for _, player_row in applicable_players.iterrows():
            global id_sum
            id_sum +=1
            try:
                player_type = player_row["player_type"]
                player_prompt_from_player = player_row["player_prompt"]
                yp_name= f"{yp_name1}_{player_type}"
                current_question = first_sentence
                if initiator == "yp":  
                    if aaa==True:
                        answer = current_question
                        history_str = f"{yp_name}：{answer}"
                        gpt_input_prompt = generate_gpt_prompt(gpt_prompt_template, yp_name, player_prompt_from_player,
                                                               user_nameprompt, history_str)
                        gpt_input_prompt = gpt_input_prompt.replace(f'_{user_nametype1}', '')
                        current_question = call_gpt_api(gpt_input_prompt)  # 1111111111111
                        current_question = current_question.replace('\n', '')  # 1111111111111
                        answer = answer.replace('\n', '')
                        current_qa_dict = {
                            "current_qa": f"{yp_name}：{answer}\nuser_name：{current_question}"  # 1111111111
                        }
                        print("--------current_qa_dict--------------", current_qa_dict)
                        update_history_recent(yp_name, current_qa_dict)
                        update_history_full(yp_name, current_qa_dict)
                        对话上文 = get_recent_history_str(yp_name, 20) + f"\n{yp_name}：{answer}"
                        if first_staus==False:
                            output_obj = {
                                "id": f"{序号}_{yp_name}_{round_num}",
                                "三级分类": 三级分类,
                                "轮次": 对话轮次,
                                "对话形式": 对话形式,
                                "对话上文": 对话上文,
                                "问题": answer_star,
                                "输出": answer,
                                "时间": get_current_time(),
                                "player": player_type,
                                "yp": yp_name,
                                "gpt完整prompt": gpt_input_prompt,
                                # "doubao": current_question,
                                # "gpt": answer
                            }
                            # print(output_obj)
                            append_jsonl(jsonl_file_path, output_obj)
                        aaa = False
                    else:
                        first_staus = False
                        answer11 = answer

                        current_question = call_doubao_result(answer, numbers, page)
                        print("----------------------", answer,current_question)

                        history_str=get_recent_history_str_full(yp_name, 10)+f"\n{yp_name}：{current_question}"
                        gpt_input_prompt = generate_gpt_prompt(gpt_prompt_template, yp_name,
                                                               player_prompt_from_player, user_nameprompt, history_str)
                        gpt_input_prompt = gpt_input_prompt.replace(f'_{user_nametype1}', '')
                        # print("历史记录：",history_str)


                        对话上文 = get_recent_history_str(yp_name, 20) + f"\n{yp_name}：{current_question}"

                        output_obj = {
                            "id": f"{序号}_{yp_name}_{round_num}",
                            "三级分类": 三级分类,
                            "轮次": 对话轮次,
                            "对话形式": 对话形式,
                            "对话上文": 对话上文,
                            "问题": answer11,
                            "输出": current_question,
                            "时间": get_current_time(),
                            "player": player_type,
                            "yp": yp_name,
                            "gpt完整prompt": gpt_input_prompt,
                            # "doubao": current_question,
                            # "gpt": answer
                        }
                        append_jsonl(jsonl_file_path, output_obj)
                        # print("========================round_num",round_num,qa_num)
                        if round_num ==qa_num:
                            current_qa_dict = {
                                "current_qa": f"{yp_name}：{current_question}"
                            }
                        else:
                            answer = call_gpt_api(gpt_input_prompt)

                            answer = answer.replace('\n', '')
                            current_qa_dict = {
                                "current_qa": f"{yp_name}：{current_question}\nuser_name：{answer}"
                            }
                        update_history_recent(yp_name, current_qa_dict)
                        update_history_full(yp_name,current_qa_dict)
                        answer_star=answer
                else:#QAQ
                    if aaa==True:
                        # print("=====================", current_question)
                        answer = call_doubao_result(current_question, numbers, page)

                        current_qa_dict = {
                            "current_qa": f"user_name：{current_question}\n{yp_name}：{answer}"
                        }
                        print("---------current_qa_dict-------------", current_qa_dict)
                        update_history_recent(yp_name, current_qa_dict)
                        update_history_full(yp_name,current_qa_dict)
                        # if first_staus == False:
                        output_obj = {
                            "id": f"{序号}_{yp_name}_{round_num}",
                            "三级分类": 三级分类,
                            "轮次": 对话轮次,
                            "对话形式": 对话形式,
                            "对话上文": get_recent_history_str(yp_name, 20),#get_recent_history_str(yp_name, 20),
                            "问题": current_question,  # user_name发起question
                            "输出":answer,
                            "时间": get_current_time(),
                            "player": player_type,
                            "yp": yp_name,
                            "gpt完整prompt": "",
                            # "doubao": answer,
                            # "gpt": current_question
                        }
                        # print(output_obj)
                        append_jsonl(jsonl_file_path, output_obj)
                        aaa = False
                    else:
                        history_str=get_recent_history_str_full(yp_name, 10)+f"\n{yp_name}：{answer}"
                        gpt_input_prompt = generate_gpt_prompt(gpt_prompt_template, yp_name, player_prompt_from_player, user_nameprompt, history_str)
                        gpt_input_prompt = gpt_input_prompt.replace(f'_{user_nametype1}', '')
                        current_question = call_gpt_api(gpt_input_prompt)
                        # print("=====================", current_question)
                        answer = call_doubao_result(current_question, numbers, page)
                        current_qa_dict = {
                            "current_qa": f"user_name：{current_question}\n{yp_name}：{answer}"
                        }
                        print("============current_qa_dict=========", current_qa_dict)
                        update_history_recent(yp_name, current_qa_dict)
                        update_history_full(yp_name,current_qa_dict)
                        # print("历史记录：", history_str)
                        output_obj = {
                            "id": f"{序号}_{yp_name}_{round_num}",
                            "三级分类": 三级分类,
                            "轮次": 对话轮次,
                            "对话形式": 对话形式,
                            "对话上文": get_recent_history_str(yp_name, 20),
                            "问题": current_question,  # user_name发起question
                            "输出": answer,
                            "时间": get_current_time(),
                            "player": player_type,
                            "yp": yp_name,
                            "gpt完整prompt": gpt_input_prompt
                        }
                        append_jsonl(jsonl_file_path, output_obj)

            except Exception as e:
                print("---------error--------:",e)
                output_obj = {
                    "id": f"{序号}_{yp_name}_{round_num}",
                    "三级分类": 三级分类,
                    "轮次": 对话轮次,
                    "对话形式": 对话形式,
                    "对话上文": "",
                    "问题": first_sentence,
                    "输出":  str(e),
                    "历史对话": "",
                    "时间": get_current_time(),
                    "player": "",
                    "yp": yp_name,
                    "gpt完整prompt": ""
                }
                append_jsonl(jsonl_file_path, output_obj)


def handle_fixed_questions(fixed_df, jsonl_file_path,page,numbers):
    for _, row in fixed_df.iterrows():
        编号 = row["编号"]
        问题 = row["问题"]
        三级分类 = row["三级分类"]
        参考信息 = row["参考信息"]
        对话形式 = "fixed题目不对话"
        yp_name = row["multiypname"]
        global id_sum
        id_sum+=1
        answer= call_doubao_result(问题, numbers, page)

        answer = answer.replace('\n', '')
        current_qa_dict = {
            "current_qa": f"{问题}\n{yp_name}：{answer}"
        }
        update_history_full(yp_name, current_qa_dict)
        update_history_recent(yp_name, current_qa_dict)
        print("----",current_qa_dict)
        # print("历史记录：",get_full_history_str(yp_name))
        output_obj = {
            "id":  f"G{编号}_{yp_name}_{id_sum}",
            "三级分类": 三级分类,
            "轮次": "",
            "对话形式": 对话形式,
            "对话上文": get_recent_history_str(yp_name, 20),
            "问题": 问题.replace("user_name：","").replace("user_name:",""),
            "输出": answer,
            "时间": get_current_time(),
            "player": "",
            "yp": yp_name,
            "gpt完整prompt": ""
        }
        append_jsonl(jsonl_file_path, output_obj)

def main(page,yp_name,version,url, state_file, headless,excel_name):
    output_folder = os.path.join(base_path, 'output')
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    timestamp = str(int(time.time()))
    file_name=f'{yp_name}_{version}_{timestamp}'
    jsonl_file_path = os.path.join(output_folder, f'{file_name}.jsonl')
    print(f"文件将保存为: {jsonl_file_path}")

    sheets = read_excel(excel_name)
    from openpyxl import load_workbook

    wb = load_workbook(excel_name)

    ws = wb['gpt_prompt']

    gpt_prompt_template = ws['B3'].value
    # print(gpt_prompt_template)
    player_df = sheets["player"]
    evaluation_df = sheets["测评分类表"]
    fixed_df = sheets["fixed题目测试集"]
    time.sleep(1)

    filtered_df = evaluation_df[evaluation_df["对话形式"] != "fixed题目"].reset_index(drop=True)

    for play_type, group in filtered_df.groupby("user_nametype1"):
        print(play_type)
        clear_chat(page)
        time.sleep(1)
        for index, row in group.iterrows():
            numbers = 2
            handle_conversation_row(row, gpt_prompt_template, player_df, jsonl_file_path, play_type, index, page,
                                    numbers)

    clear_chat(page)
    time.sleep(1)
    numbers=2
    # handle_fixed_questions(fixed_df, jsonl_file_path,page,numbers)

    records = []
    with open(jsonl_file_path, "r", encoding="utf-8") as f:
        for line in f:
            records.append(json.loads(line))


    df_output = pd.DataFrame(records)
    output_filename = os.path.join(output_folder, f'{file_name}.xlsx')
    with pd.ExcelWriter(output_filename) as writer:
        df_output.to_excel(writer, sheet_name='原始数据', index=False)

    df = pd.read_excel(output_filename, sheet_name='原始数据')

    c_data = df[df['id'].str.startswith('N')]
    non_c_data = df[~df['id'].str.startswith('N')]

    def extract_last_per_group(group):
        return group.iloc[-1]

    c_data_processed = c_data.groupby(c_data['id'].str.split('_').str[0]).apply(extract_last_per_group).reset_index(
        drop=True)

    final_df = pd.concat([c_data_processed, non_c_data], ignore_index=True)

    with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        final_df.to_excel(writer, sheet_name='后处理数据', index=False)
    print(f"文件已保存为: {output_filename}")

if __name__ == "__main__":
    with sync_playwright() as playwright:
        yp_name=""

        version="iiiiiii" 


        excel_name = 'iiu.xlsx' 
        url = "" 
        state_file = 'hhh.json' 

        print(state_file)

        headless=False
        page,context,browser = doubao_star(playwright, url, state_file, headless)
        main(page,yp_name,version,url, state_file, headless,excel_name)
        doubao_end(page,context,browser)


















